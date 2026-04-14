"""
人事評価システム Phase 1
等級スキルシート + 目標管理評価シート の評価支援ツール
"""
import re
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
import zipfile

st.set_page_config(
    page_title="人事評価システム",
    page_icon="🏢",
    layout="wide",
)

# ══════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════

GRADE_TABLE = [
    (78, "Ⅴ-3"), (74, "Ⅴ-2"), (70, "Ⅴ-1"),
    (66, "Ⅳ-3"), (63, "Ⅳ-2"), (60, "Ⅳ-1"),
    (50, "Ⅲ-3"), (45, "Ⅲ-2"), (40, "Ⅲ-1"),
    (35, "Ⅱ-3"), (30, "Ⅱ-2"), (25, "Ⅱ-1"),
    (20, "Ⅰ-3"), (15, "Ⅰ-2"), (10, "Ⅰ-1"),
]

WEIGHT_TABLE = {
    ("営業", "Ⅰ"): [4, 4, 1, 1],
    ("営業", "Ⅱ"): [5, 3, 1, 1],
    ("営業", "Ⅲ"): [6, 2, 1, 1],
    ("営業", "Ⅳ"): [6, 1, 2, 1],
    ("営業", "Ⅴ"): [6, 1, 2, 1],
    ("非営業", "Ⅰ"): [1, 3, 3, 3],
    ("非営業", "Ⅱ"): [2, 3, 2, 3],
    ("非営業", "Ⅲ"): [3, 3, 2, 2],
    ("非営業", "Ⅳ"): [3, 4, 2, 1],
    ("非営業", "Ⅴ"): [4, 4, 1, 1],
}

EVAL_COEFF = {"S": 1.5, "A": 1.2, "B": 1.0, "C": 0.8, "D": 0.0, "": 0.0, "-": 0.0}

SKILL_ITEMS = [
    ("①プロジェクト成功スキル",              38, 2, "必須"),
    ("②課題発見・解決スキル",               44, 2, "必須"),
    ("③グロースさせるスキル",               50, 2, "必須"),
    ("④WEBサイト設計スキル",               56, 2, "必須"),
    ("⑤コミュニケーション・ファシリテーション", 62, 2, "必須"),
    ("⑥テクニカルスキル",                  68, 1, "選択"),
    ("⑦市場調査・データ分析",               74, 1, "選択"),
    ("⑧マーケティングコミュニケーション",      80, 1, "選択"),
    ("⑨業務改善",                          86, 1, "選択"),
]

CF_ITEMS = [
    ("①PM実績",         95,  1, "PM"),
    ("②PMスキル",       101, 1, "PM"),
    ("③面談・評価業務",  107, 1, "エバリュエイター"),
    ("④マネジメント実績", 113, 1, "エバリュエイター"),
    ("⑤アドバイザリー",  119, 1, "プロフェッショナルアドバイザー"),
    ("⑥トレーナー業務",  125, 1, "トレーナー"),
    ("⑦部門目標意識",   131, 2, "目標達成マインド"),
]

TARGET_AXES = ["①業績貢献", "②業績プロセス", "③会社貢献", "④パーソナリティ"]
TARGET_ROWS = {"①業績貢献": 12, "②業績プロセス": 19, "③会社貢献": 24, "④パーソナリティ": 29}

COL_UL_COMMENT_LOWER = 18
COL_SELF_LOWER       = 19
COL_UL_LOWER         = 20
COL_FINAL_LOWER      = 21
COL_UL_TOTAL         = 22
COL_FINAL_TOTAL      = 23
COL_WEIGHT           = 24
COL_SCORE            = 25
COL_SKILL_SCORE      = 7
COL_SKILL_TOTAL      = 8
COL_CRITERIA         = 5   # 判断基準テキスト列

# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════

def score_to_grade(total: float) -> str:
    for threshold, grade in GRADE_TABLE:
        if total >= threshold:
            return grade
    return "Ⅰ-1"

def grade_roman(grade_str: str) -> str:
    for r in ["Ⅴ", "Ⅳ", "Ⅲ", "Ⅱ", "Ⅰ"]:
        if str(grade_str or "").startswith(r):
            return r
    return "Ⅱ"

def get_weights(job_type: str, grade_str: str) -> list:
    roman = grade_roman(grade_str)
    return WEIGHT_TABLE.get((job_type, roman), [3, 3, 2, 2])

def calc_skill_totals(skill_scores: dict, cf_scores: dict) -> tuple:
    s = sum(v["score"] * v["coeff"] for v in skill_scores.values())
    c = sum(v["score"] * v["coeff"] for v in cf_scores.values())
    return s, c

def safe_int(val, default=0) -> int:
    try:
        return int(float(val)) if val is not None else default
    except:
        return default

def safe_str(val, default="") -> str:
    return str(val).strip() if val not in (None, "") else default

def extract_name_from_filename(filename: str) -> str:
    """ファイル名から氏名を推定する（末尾の8桁日付を除いた最後のセグメント）"""
    name = filename.replace(".xlsx", "").replace(".xls", "")
    name = re.sub(r"_\d{8}$", "", name)
    parts = [p.strip() for p in name.split("_") if p.strip()]
    return parts[-1] if parts else ""

# ══════════════════════════════════════════════════════════════
# PARSERS
# ══════════════════════════════════════════════════════════════

def _detect_tab_priority(sheet_name: str) -> int:
    """タブ名の接尾語から優先度を返す（低いほど優先）。"""
    if "_自己評価" in sheet_name: return 3
    if "_UL評価"   in sheet_name: return 2
    if "_最終評価" in sheet_name: return 1
    return 0  # 接尾語なし = 最優先

def _extract_yyyymm(sheet_name: str) -> int:
    m = re.search(r"(\d{6})", sheet_name)
    return int(m.group(1)) if m else 0

def select_skill_sheets(wb) -> tuple:
    """
    等級要件シートを抽出し、同一年月の場合は優先度で選択する。
    優先度: なにもなし > 最終評価 > UL評価 > 自己評価
    Returns: (ws_curr, ws_prev)
    """
    skill_sheets = []
    for s in wb.sheetnames:
        try:
            val = str(wb[s].cell(2, 2).value or "")
            if "等級要件" in val:
                yyyymm   = _extract_yyyymm(s)
                priority = _detect_tab_priority(s)
                skill_sheets.append((yyyymm, priority, s))
        except:
            pass

    if not skill_sheets:
        return None, None

    # 同一年月ごとにグループ化し、最優先タブを選択
    groups: dict = {}
    for yyyymm, priority, name in skill_sheets:
        groups.setdefault(yyyymm, []).append((priority, name))

    selected = []
    for yyyymm, candidates in groups.items():
        best = sorted(candidates, key=lambda x: x[0])[0]
        selected.append((yyyymm, best[1]))

    # 年月で降順ソート（最新=今期）
    selected = sorted(selected, key=lambda x: x[0], reverse=True)

    ws_curr = wb[selected[0][1]] if selected            else None
    ws_prev = wb[selected[1][1]] if len(selected) >= 2 else None
    return ws_curr, ws_prev

def read_skill_ws_dynamic(ws) -> tuple:
    """
    動的にスキル項目・CF項目を検出する。
    【スキルシート】【クロスファンクショナルシート】の区切りを使い、
    ①〜⑨の丸数字で始まる行を項目ヘッダーとして認識する。
    Returns: (skill_dict, cf_dict, skill_criteria, cf_criteria)
    """
    MARU = set("①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮")

    def is_item_header(val) -> bool:
        s = str(val or "").strip()
        return bool(s) and s[0] in MARU

    def get_coeff(val) -> int:
        m = re.search(r"係数[×x×](\d+)", str(val or ""))
        return int(m.group(1)) if m else 1

    # セクション区切りを検出
    skill_start = cf_start = end_row = None
    for r in range(1, 250):
        for col in [1, 2]:
            v = str(ws.cell(r, col).value or "")
            if "【スキルシート】" in v and skill_start is None:
                skill_start = r + 1
            if "【クロスファンクショナルシート】" in v and cf_start is None:
                cf_start = r + 1
            if ("【加点" in v or "【減点" in v) and cf_start is not None and end_row is None:
                end_row = r

    if skill_start is None: skill_start = 36
    if cf_start    is None: cf_start    = 80
    if end_row     is None: end_row     = 200

    def parse_section(start: int, end: int) -> tuple:
        items: dict    = {}
        criteria: dict = {}
        cur_name  = None
        cur_coeff = 1
        cur_score = 0
        cur_crit: dict = {}

        for r in range(start, end):
            c5 = ws.cell(r, 5).value
            c7 = ws.cell(r, 7).value
            c8 = ws.cell(r, 8).value

            if is_item_header(c5):
                item_name = str(c5).split("\n")[0].strip()
                if cur_name and cur_name != item_name and cur_name not in items:
                    items[cur_name]    = {"score": cur_score, "coeff": cur_coeff, "kind": ""}
                    criteria[cur_name] = cur_crit
                if item_name not in items:
                    cur_name  = item_name
                    cur_score = safe_int(c7)
                    cur_coeff = get_coeff(c8)
                    cur_crit  = {}
            elif cur_name and c5 and isinstance(c7, (int, float)) and 1 <= float(c7) <= 5:
                cur_crit[int(float(c7))] = str(c5)[:100]

        if cur_name and cur_name not in items:
            items[cur_name]    = {"score": cur_score, "coeff": cur_coeff, "kind": ""}
            criteria[cur_name] = cur_crit

        return items, criteria

    skill_items, skill_crit = parse_section(skill_start, cf_start)
    cf_items,    cf_crit    = parse_section(cf_start, end_row)
    return skill_items, cf_items, skill_crit, cf_crit

def parse_skill_file(file_bytes: bytes, filename: str = "") -> dict:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    ws_curr, ws_prev = select_skill_sheets(wb)

    # フォールバック
    if ws_curr is None:
        sheets  = wb.sheetnames
        ws_curr = wb[sheets[0]]
        ws_prev = wb[sheets[1]] if len(sheets) > 1 else None

    curr_skill, curr_cf, skill_crit, cf_crit = read_skill_ws_dynamic(ws_curr)
    if ws_prev:
        prev_skill, prev_cf, _, _ = read_skill_ws_dynamic(ws_prev)
    else:
        prev_skill, prev_cf = {}, {}

    s = sum(v["score"] * v["coeff"] for v in curr_skill.values())
    c = sum(v["score"] * v["coeff"] for v in curr_cf.values())

    deduct = safe_int(ws_curr.cell(4, 11).value)
    total  = s + c - deduct

    name         = safe_str(ws_curr.cell(4, 3).value)
    job_type_raw = safe_str(ws_curr.cell(6, 3).value)

    name_from_file = False
    if not name and filename:
        name = extract_name_from_filename(filename)
        name_from_file = True

    return {
        "name":            name,
        "name_from_file":  name_from_file,
        "period":          safe_str(ws_curr.cell(4, 7).value),
        "department":      safe_str(ws_curr.cell(5, 3).value),
        "job_type_raw":    job_type_raw,
        "job_type_label":  job_type_raw or "未設定",
        "job_type":        "非営業",
        "skill_scores":    curr_skill,
        "cf_scores":       curr_cf,
        "prev_skill":      prev_skill,
        "prev_cf":         prev_cf,
        "skill_criteria":  skill_crit,
        "cf_criteria":     cf_crit,
        "skill_total":     s,
        "cf_total":        c,
        "deduct":          deduct,
        "total":           total,
        "grade":           score_to_grade(total),
        "prev_grade":      safe_str(ws_curr.cell(5, 12).value),
        "change_comments": {},
        "file_bytes":      file_bytes,
    }


def parse_target_file(file_bytes: bytes, filename: str = "") -> dict:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    axes = {}
    for axis in TARGET_AXES:
        row = TARGET_ROWS[axis]
        axes[axis] = {
            # ── 上期
            "upper_jisseki":   safe_str(ws.cell(row, 4).value),   # 実績・成果
            "upper_target":    safe_str(ws.cell(row, 6).value),   # 目標設定
            "upper_self_rev":  safe_str(ws.cell(row, 7).value),   # 本人振り返り
            "upper_ul_cmt":    safe_str(ws.cell(row, 8).value),   # ULコメント
            "upper_self_eval": safe_str(ws.cell(row, 9).value),   # 本人評価
            "upper_ul_eval":   safe_str(ws.cell(row, 10).value),  # UL評価
            "upper_final":     safe_str(ws.cell(row, 11).value),  # 最終評価
            "upper_weight":    ws.cell(row, 12).value,             # ウェイト
            "upper_score":     ws.cell(row, 13).value,             # 獲得スコア
            # ── 下期
            "lower_jisseki":   safe_str(ws.cell(row, 14).value),  # 実績・成果
            "lower_target":    safe_str(ws.cell(row, 16).value),  # 目標設定
            "lower_self_rev":  safe_str(ws.cell(row, 17).value),  # 本人振り返り
            "ul_comment":      safe_str(ws.cell(row, COL_UL_COMMENT_LOWER).value),
            "self_eval":       safe_str(ws.cell(row, COL_SELF_LOWER).value),
            "ul_eval":         safe_str(ws.cell(row, COL_UL_LOWER).value),
            "final_eval":      safe_str(ws.cell(row, COL_FINAL_LOWER).value),
            # ── 通期
            "ul_total":        safe_str(ws.cell(row, COL_UL_TOTAL).value),
            "final_total":     safe_str(ws.cell(row, COL_FINAL_TOTAL).value),
            "weight":          ws.cell(row, COL_WEIGHT).value,
            "score":           ws.cell(row, COL_SCORE).value,
        }

    # 上期総合評価（行34）
    upper_summary = {
        "self_eval":  safe_str(ws.cell(34, 9).value),
        "ul_eval":    safe_str(ws.cell(34, 10).value),
        "final_eval": safe_str(ws.cell(34, 11).value),
        "weight":     ws.cell(34, 12).value,
        "score":      ws.cell(34, 13).value,
    }

    name = safe_str(ws.cell(4, 4).value)
    name_from_file = False
    if not name and filename:
        name = extract_name_from_filename(filename)
        name_from_file = True

    return {
        "name":           name,
        "name_from_file": name_from_file,
        "period":         safe_str(ws.cell(2, 2).value),
        "department":     safe_str(ws.cell(5, 4).value),
        "grade_str":      safe_str(ws.cell(7, 4).value),
        "job_type":       "非営業",
        "axes":           axes,
        "upper_summary":  upper_summary,
        "file_bytes":     file_bytes,
    }

def is_skill_file(wb) -> bool:
    try:
        val = str(wb[wb.sheetnames[0]].cell(2, 2).value or "")
        return "等級要件" in val
    except:
        return False

# ══════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════

for _k, _v in [("skill_members", {}), ("target_members", {}),
               ("audit_log", []), ("skill_ver", 0), ("cf_ver", 0)]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

def audit(member, kind, field, prev, new, comment=""):
    st.session_state.audit_log.append({
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "member":    member,
        "kind":      kind,
        "field":     field,
        "prev":      str(prev),
        "new":       str(new),
        "comment":   comment,
    })

# ══════════════════════════════════════════════════════════════
# EXPORT HELPERS
# ══════════════════════════════════════════════════════════════

RED_FONT    = Font(color="FF0000", bold=True)
GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)

def export_skill_excel(member_data: dict) -> bytes:
    d  = member_data
    wb = load_workbook(BytesIO(d["file_bytes"]))
    ws = wb[wb.sheetnames[0]]

    for name, row, coeff, _ in SKILL_ITEMS:
        new_score  = d["skill_scores"][name]["score"]
        prev_score = d["prev_skill"].get(name, {}).get("score") if d["prev_skill"] else None
        ws.cell(row, COL_SKILL_SCORE).value = new_score
        ws.cell(row, COL_SKILL_TOTAL).value = new_score * coeff
        if prev_score is not None and new_score != prev_score:
            ws.cell(row, COL_SKILL_SCORE).font = RED_FONT
            ws.cell(row, COL_SKILL_TOTAL).font = RED_FONT

    for name, row, coeff, _ in CF_ITEMS:
        new_score  = d["cf_scores"][name]["score"]
        prev_score = d["prev_cf"].get(name, {}).get("score") if d["prev_cf"] else None
        ws.cell(row, COL_SKILL_SCORE).value = new_score
        ws.cell(row, COL_SKILL_TOTAL).value = new_score * coeff
        if prev_score is not None and new_score != prev_score:
            ws.cell(row, COL_SKILL_SCORE).font = RED_FONT
            ws.cell(row, COL_SKILL_TOTAL).font = RED_FONT

    ws.cell(9,  2).value = d["skill_total"]
    ws.cell(25, 2).value = d["cf_total"]

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def export_target_excel(member_data: dict) -> bytes:
    d  = member_data
    wb = load_workbook(BytesIO(d["file_bytes"]))
    ws = wb.active

    ws.cell(7, 4).value = d.get("grade_str", "")

    # 各評価軸（行12/19/24/29）への書き込み
    for axis in TARGET_AXES:
        row = TARGET_ROWS[axis]
        ax  = d["axes"][axis]
        ws.cell(row, COL_UL_COMMENT_LOWER).value = ax.get("ul_comment", "")
        ws.cell(row, COL_UL_LOWER).value          = ax.get("ul_eval", "")
        ws.cell(row, COL_FINAL_LOWER).value        = ax.get("final_eval", "")
        ws.cell(row, COL_UL_TOTAL).value           = ax.get("ul_total", "")    # V列
        ws.cell(row, COL_FINAL_TOTAL).value        = ax.get("final_total", "") # W列
        ws.cell(row, COL_WEIGHT).value             = ax.get("weight")
        ws.cell(row, COL_SCORE).value              = ax.get("score")

    # 行34（総合評価行）への書き込み
    # 通期スコア合計・ウェイト合計・目安評価を算出
    grade_str = d.get("grade_str", "")
    job_type  = d.get("job_type", "非営業")
    weights   = get_weights(job_type, grade_str)
    weight_sum = sum(weights)

    total_annual = sum(
        EVAL_COEFF.get(
            d["axes"][ax].get("final_total", "") or
            d["axes"][ax].get("ul_total", ""), 0.0
        ) * weights[i]
        for i, ax in enumerate(TARGET_AXES)
    )
    rank_annual = score_to_rank(total_annual, weight_sum)
    rate_annual = round(total_annual / weight_sum, 4) if weight_sum > 0 else 0

    ws.cell(34, COL_UL_TOTAL).value    = rank_annual   # V34: 通期UL評価
    ws.cell(34, COL_FINAL_TOTAL).value = rank_annual   # W34: 通期最終評価
    ws.cell(34, COL_WEIGHT).value      = weight_sum    # X34: ウェイト合計
    ws.cell(34, COL_SCORE).value       = round(total_annual, 2)  # Y34: 獲得スコア
    ws.cell(34, 26).value              = rank_annual   # Z34: 最終評価

    # 達成率（行35のY列相当）
    ws.cell(35, COL_SCORE).value = rate_annual

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def export_summary_excel() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "評価一覧"

    headers = [
        "No.", "氏名", "職種", "職種区分", "等級",
        "スキルスコア", "CFスコア", "合計スコア",
        "①業績貢献 UL評価", "②業績プロセス UL評価",
        "③会社貢献 UL評価", "④パーソナリティ UL評価",
        "目標管理スコア",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    all_names = sorted(set(
        list(st.session_state.skill_members.keys()) +
        list(st.session_state.target_members.keys())
    ))

    for i, nm in enumerate(all_names, 1):
        row = [i, nm]
        if nm in st.session_state.skill_members:
            sd = st.session_state.skill_members[nm]
            row += [sd["job_type_label"], sd["job_type"], sd["grade"],
                    sd["skill_total"], sd["cf_total"], sd["total"]]
        else:
            row += ["", "", "", "", "", ""]

        if nm in st.session_state.target_members:
            td = st.session_state.target_members[nm]
            evals = [td["axes"][ax].get("ul_eval", "") for ax in TARGET_AXES]
            total_score = sum(
                EVAL_COEFF.get(td["axes"][ax].get("ul_eval", ""), 0) *
                (td["axes"][ax].get("weight") or 0)
                for ax in TARGET_AXES
            )
            row += evals + [round(total_score, 2)]
        else:
            row += ["", "", "", "", ""]

        ws.append(row)
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = GRAY_FILL

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════
# UI: スキル評価カード（判断基準インライン表示）
# ══════════════════════════════════════════════════════════════

def _criteria_html(pt: int, text: str, matched: list, current_score: int) -> str:
    """1点分の判断基準HTMLカードを生成する。"""
    is_matched = bool(matched)
    is_current = (pt == current_score)
    border = "#4CAF50" if is_matched else ("#90CAF9" if is_current else "#e0e0e0")
    bg     = "#f0faf0" if is_matched else ("#e8f4fd" if is_current else "#fafafa")
    badges = " ".join(
        f'<span style="background:#4CAF50;color:#fff;border-radius:3px;'
        f'padding:1px 7px;font-size:0.78em;margin-left:4px;">{m}</span>'
        for m in matched
    )
    return (
        f'<div style="background:{bg};border-left:4px solid {border};'
        f'border-radius:0 5px 5px 0;padding:7px 12px;margin:4px 0;">'
        f'<strong style="font-size:0.95em;">{pt}点</strong>{badges}<br>'
        f'<span style="font-size:0.87em;color:#333;line-height:1.5;">{text}</span>'
        f'</div>'
    )


def render_skill_comparison(
    items: list,
    score_key: str,
    prev_key: str,
    criteria_key: str,
    ver_key: str,
    filter_members: list = None,
):
    """
    スキル評価カード型UI。
    各項目に判断基準（▼クリックで展開）をインライン表示する。
    """
    all_members = list(st.session_state.skill_members.keys())
    members     = filter_members if filter_members is not None else all_members

    if not all_members:
        st.info("スキルシートをアップロードしてください。")
        return
    if not members:
        st.info("この職種のメンバーはいません。")
        return

    SCORE_OPTS = [0, 1, 2, 3, 4, 5]
    BADGE_COLOR = {"必須": "#4472C4", "選択": "#ED7D31"}

    # 判断基準は最初のメンバーのExcelから取得（全員同じ構造想定）
    all_criteria = st.session_state.skill_members[members[0]].get(criteria_key, {})

    st.caption(
        "📌 各項目のプルダウンで今期スコアを選択 → 下の「💾 保存」ボタンを押して確定。"
        "「判断基準」をクリックすると点数の意味が確認できます。"
    )

    # ── 項目カードを1行ずつ描画 ─────────────────────────
    for item_name, _excel_row, coeff, category in items:
        criteria = all_criteria.get(item_name, {})
        badge_c  = BADGE_COLOR.get(category, "#888")

        # 各メンバーの現スコア（表示用・判断基準ハイライト用）
        cur_scores = {
            m: st.session_state.skill_members[m][score_key]
               .get(item_name, {}).get("score", 0)
            for m in members
        }

        # カードヘッダー
        st.markdown(
            f'<div style="background:#f0f2f6;border-left:5px solid {badge_c};'
            f'padding:7px 14px;border-radius:0 6px 0 0;margin-top:14px;">'
            f'<strong>{item_name}</strong>'
            f'<span style="background:{badge_c};color:#fff;border-radius:3px;'
            f'padding:1px 8px;font-size:0.78em;margin-left:8px;">{category}</span>'
            f'<span style="color:#777;font-size:0.83em;margin-left:8px;">係数×{coeff}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # スコア入力行：メンバーごとに「前期N → 今期[▼]」を横並び
        score_cols = st.columns(len(members))
        for col_i, m in enumerate(members):
            d      = st.session_state.skill_members[m]
            prev_d = d.get(prev_key, {})
            prev   = prev_d.get(item_name, {}).get("score", None) if prev_d else None
            curr   = int(d[score_key].get(item_name, {}).get("score", 0))
            prev_label = f"（前期: {prev}点）" if prev is not None else ""

            with score_cols[col_i]:
                st.selectbox(
                    f"**{m}** {prev_label}",
                    options=SCORE_OPTS,
                    index=curr,
                    key=f"sb_{ver_key}_{m}_{item_name}",
                )

        # 判断基準（インライン折りたたみ）
        if criteria:
            with st.expander("　📋 判断基準を確認する", expanded=False):
                for pt in [5, 4, 3, 2, 1]:
                    text = criteria.get(pt)
                    if not text:
                        continue
                    matched = [m for m, s in cur_scores.items() if s == pt]
                    # 現在のスコアはどのメンバーでもよい（代表として最初のメンバー）
                    st.markdown(
                        _criteria_html(pt, text, matched, cur_scores.get(members[0], 0)),
                        unsafe_allow_html=True,
                    )

    # ── 保存ボタン ──────────────────────────────────────
    st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
    st.info(
        "💡 **スコアを変更した場合：** 保存ボタンを押すと"
        "**変更理由のコメント入力欄**が表示されます（任意）。"
        "コメントは「📜 評価履歴」タブで確認できます。"
    )
    col_btn, col_msg = st.columns([1, 4])

    if col_btn.button("💾 スコアを保存", key=f"save_btn_{ver_key}", type="primary"):
        changes = []
        for item_name, _, _, _ in items:
            for m in members:
                old     = int(st.session_state.skill_members[m][score_key]
                              .get(item_name, {}).get("score", 0))
                new_val = int(st.session_state.get(f"sb_{ver_key}_{m}_{item_name}", old))
                if old != new_val:
                    changes.append((m, item_name, old, new_val))
                    st.session_state.skill_members[m][score_key][item_name]["score"] = new_val

        # 全員のスコアを再計算
        for m in all_members:
            d = st.session_state.skill_members[m]
            s, c = calc_skill_totals(d["skill_scores"], d["cf_scores"])
            d["skill_total"] = s
            d["cf_total"]    = c
            d["total"]       = s + c - d.get("deduct", 0)
            d["grade"]       = score_to_grade(d["total"])

        if changes:
            st.session_state[f"pending_{ver_key}"] = changes
        else:
            col_msg.info("変更はありませんでした。")

        cur = st.session_state.get(ver_key, 0)
        st.session_state[ver_key] = cur + 1
        st.rerun()

    # ── 変更コメント入力 ────────────────────────────────
    pending_key = f"pending_{ver_key}"
    if pending_key in st.session_state and st.session_state[pending_key]:
        changes = st.session_state[pending_key]
        st.success(f"✅ {len(changes)}件のスコアを更新しました。変更理由があれば入力してください（任意）。")

        comment_vals = {}
        for m, field, old, new_val in changes:
            arrow = "⬆️" if new_val > old else "⬇️"
            comment_vals[f"{m}_{field}"] = st.text_input(
                f"{arrow} **{m}** ｜ {field}　{old}点 → {new_val}点",
                key=f"cmt_{ver_key}_{m}_{field}",
                placeholder="変更理由・コメント（任意・入力しなくてもOK）",
            )

        if st.button("✅ 確定して履歴に保存", key=f"confirm_{ver_key}"):
            for m, field, old, new_val in changes:
                comment = comment_vals.get(f"{m}_{field}", "")
                audit(m, "スキル", field, old, new_val, comment)
                if comment:
                    st.session_state.skill_members[m]["change_comments"][field] = {
                        "prev": old, "new": new_val, "comment": comment,
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    }
            del st.session_state[pending_key]
            st.rerun()

    # ── スコアサマリー ──────────────────────────────────
    st.divider()
    st.markdown("#### 📊 スコアサマリー（全項目集計）")
    summary = {}
    for m in members:
        d = st.session_state.skill_members[m]
        # スキルスコア・CFスコアを正しく表示（calc_skill_totalsで確認）
        sk = sum(v["score"] * v["coeff"] for v in d["skill_scores"].values())
        cf = sum(v["score"] * v["coeff"] for v in d["cf_scores"].values())
        total = sk + cf - d.get("deduct", 0)
        grade = score_to_grade(total)
        # セッション値も更新（整合性確保）
        d["skill_total"] = sk
        d["cf_total"]    = cf
        d["total"]       = total
        d["grade"]       = grade
        summary[m] = {
            "職種":           d["job_type_label"],
            "スキルスコア":   sk,
            "CFスコア":       cf,
            "合計スコア":     total,
            "等級":           grade,
        }
    df_summary = pd.DataFrame(summary).T
    st.dataframe(
        df_summary.style.highlight_max(
            subset=["スキルスコア", "CFスコア", "合計スコア"], color="#d4edda"
        ),
        use_container_width=True,
    )
    st.caption("※ スキルスコア＝必須・選択項目の合計　CFスコア＝クロスファンクショナルの合計")

# ══════════════════════════════════════════════════════════════
# MAIN LAYOUT
# ══════════════════════════════════════════════════════════════

st.title("🏢 人事評価システム")

t_upload, t_skill, t_target, t_history, t_export = st.tabs([
    "📁 アップロード",
    "📊 等級スキルシート",
    "📋 目標管理評価",
    "📜 評価履歴",
    "📥 出力",
])

# ──────────────────────────────────────────────────────────────
# TAB 1: UPLOAD
# ──────────────────────────────────────────────────────────────
with t_upload:
    st.header("ファイルアップロード")
    st.info(
        "**スキルシート** または **目標管理評価シート** の Excel をドラッグ＆ドロップしてください。\n"
        "複数ファイルを一度にアップロードできます。"
    )

    uploaded = st.file_uploader(
        "Excelファイルをここにドラッグ＆ドロップ",
        type="xlsx",
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    if uploaded:
        for f in uploaded:
            try:
                b  = f.read()
                wb = load_workbook(BytesIO(b), data_only=True)

                if is_skill_file(wb):
                    d  = parse_skill_file(b, filename=f.name)
                    nm = d["name"]
                    if not nm:
                        st.error(f"❌ {f.name}: 氏名が読み取れませんでした。ファイル内の氏名セルを確認してください。")
                        continue
                    if d["name_from_file"]:
                        st.warning(
                            f"⚠️ **{nm}**: 氏名セルが空のため、ファイル名から「**{nm}**」として登録しました。"
                            f"正しくない場合は下の一覧から修正できます。"
                        )
                    if nm not in st.session_state.skill_members:
                        st.session_state.skill_members[nm] = d
                        st.success(
                            f"✅ スキルシート登録: **{nm}** "
                            f"（職種: {d['job_type_label']} ／ 等級: {d['grade']} ／ {d['total']}点）"
                        )
                    else:
                        st.info(f"ℹ️ 「{nm}」は既に登録済みです。")

                else:
                    d  = parse_target_file(b, filename=f.name)
                    nm = d["name"]
                    if not nm:
                        st.error(f"❌ {f.name}: 氏名が読み取れませんでした。ファイル内の氏名セルを確認してください。")
                        continue
                    if d["name_from_file"]:
                        st.warning(
                            f"⚠️ **{nm}**: 氏名セルが空のため、ファイル名から「**{nm}**」として登録しました。"
                            f"正しくない場合は下の一覧から修正できます。"
                        )
                    if nm not in st.session_state.target_members:
                        st.session_state.target_members[nm] = d
                        st.success(f"✅ 目標管理シート登録: **{nm}**")
                    else:
                        st.info(f"ℹ️ 「{nm}」は既に登録済みです。")

            except Exception as e:
                st.error(f"❌ {f.name} の読み込みエラー: {e}")

    st.divider()
    c1, c2 = st.columns(2)

    with c1:
        st.subheader(f"📊 スキルシート一覧 ({len(st.session_state.skill_members)}名)")
        if not st.session_state.skill_members:
            st.caption("未登録")
        for nm, d in list(st.session_state.skill_members.items()):
            with st.expander(
                f"**{nm}** ｜ {d['job_type_label']} ｜ {d['grade']} ｜ {d['total']}点",
                expanded=False,
            ):
                new_label = st.text_input(
                    "職種名（タブ切り替え用）",
                    value=d["job_type_label"],
                    key=f"jl_s_{nm}",
                    placeholder="例: ディレクター",
                )
                d["job_type_label"] = new_label or d["job_type_label"]

                jt = st.radio(
                    "ウェイト計算用区分",
                    ["営業", "非営業"],
                    key=f"jt_s_{nm}",
                    index=0 if d["job_type"] == "営業" else 1,
                    horizontal=True,
                    help="目標管理のウェイト計算に使います",
                )
                d["job_type"] = jt

                st.caption(
                    f"スキル: {d['skill_total']}点 ／ CF: {d['cf_total']}点 ／ "
                    f"減点: -{d['deduct']}点 ／ 前期等級: {d['prev_grade'] or '–'}"
                )
                if st.button("🗑 削除", key=f"del_s_{nm}"):
                    del st.session_state.skill_members[nm]
                    st.rerun()

    with c2:
        st.subheader(f"📋 目標管理シート一覧 ({len(st.session_state.target_members)}名)")
        if not st.session_state.target_members:
            st.caption("未登録")
        for nm, d in list(st.session_state.target_members.items()):
            if nm in st.session_state.skill_members:
                d["grade_str"] = st.session_state.skill_members[nm]["grade"]
                d["job_type"]  = st.session_state.skill_members[nm]["job_type"]

            with st.expander(
                f"**{nm}** ｜ 等級: {d.get('grade_str', '未設定')}",
                expanded=False,
            ):
                jt = st.radio(
                    "ウェイト計算用区分",
                    ["営業", "非営業"],
                    key=f"jt_t_{nm}",
                    index=0 if d["job_type"] == "営業" else 1,
                    horizontal=True,
                )
                d["job_type"] = jt

                if nm not in st.session_state.skill_members:
                    grade_in = st.text_input(
                        "等級（手動入力）",
                        value=d.get("grade_str", ""),
                        key=f"grade_in_{nm}",
                        placeholder="例: Ⅱ-1",
                    )
                    d["grade_str"] = grade_in

                if st.button("🗑 削除", key=f"del_t_{nm}"):
                    del st.session_state.target_members[nm]
                    st.rerun()

# ──────────────────────────────────────────────────────────────
# TAB 2: SKILL EVALUATION
# ──────────────────────────────────────────────────────────────
with t_skill:
    st.header("等級スキルシート評価")

    if not st.session_state.skill_members:
        st.info("👈 まずアップロードタブでスキルシートを登録してください。")
    else:
        # 職種ラベルの一覧を収集
        all_job_labels = sorted(set(
            d["job_type_label"]
            for d in st.session_state.skill_members.values()
            if d["job_type_label"] and d["job_type_label"] != "未設定"
        ))

def score_to_rank(total_score: float, weight_sum: float) -> str:
    """獲得スコアとウェイト合計から達成率で目安評価を計算（11段階）
    S: 135%超
    A+: 125%超〜135%以下 / A: 115%超〜125%以下 / A－: 105%超〜115%以下
    B+: 102%超〜105%以下 / B: 98%超〜102%以下 / B－: 95%超〜98%以下
    C+: 85%超〜95%以下  / C: 75%超〜85%以下  / C－: 65%超〜75%以下
    D: 65%以下
    """
    if weight_sum <= 0:
        return "–"
    rate = total_score / weight_sum
    if rate > 1.35:  return "S"
    if rate > 1.25:  return "A+"
    if rate > 1.15:  return "A"
    if rate > 1.05:  return "A－"
    if rate > 1.02:  return "B+"
    if rate > 0.98:  return "B"
    if rate > 0.95:  return "B－"
    if rate > 0.85:  return "C+"
    if rate > 0.75:  return "C"
    if rate > 0.65:  return "C－"
    return "D"

RANK_COLOR = {
    "S":  "#1565C0",
    "A+": "#1976D2", "A": "#1E88E5", "A－": "#42A5F5",
    "B+": "#2E7D32", "B": "#388E3C", "B－": "#66BB6A",
    "C+": "#E65100", "C": "#F57F17", "C－": "#FFA726",
    "D":  "#B71C1C",
    "–":  "#9E9E9E",
}

# ──────────────────────────────────────────────────────────────
# TAB 2: SKILL EVALUATION
# ──────────────────────────────────────────────────────────────
with t_skill:
    st.header("等級スキルシート評価")

    if not st.session_state.skill_members:
        st.info("👈 まずアップロードタブでスキルシートを登録してください。")
    else:
        # 部署一覧を収集（アップロード一覧で設定した部署）
        all_depts = sorted(set(
            d.get("department", "") or "未設定"
            for d in st.session_state.skill_members.values()
        ))
        # 部署が1種類 or 全員未設定ならタブ不要
        show_dept_tabs = len(all_depts) > 1 or (len(all_depts) == 1 and all_depts[0] != "未設定")

        dept_groups = {}
        for nm, d in st.session_state.skill_members.items():
            dept = d.get("department", "") or "未設定"
            dept_groups.setdefault(dept, []).append(nm)

        def render_job_tabs(dept_members: list, dept_key: str):
            """職種タブ＋スキル比較グリッドを描画"""
            all_job_labels = sorted(set(
                st.session_state.skill_members[m]["job_type_label"]
                for m in dept_members
                if st.session_state.skill_members[m]["job_type_label"] not in ("", "未設定")
            ))
            tab_labels = ["👥 全員"] + [f"🏷️ {jl}" for jl in all_job_labels]
            job_tabs = st.tabs(tab_labels)

            for tab_idx, job_tab in enumerate(job_tabs):
                with job_tab:
                    if tab_idx == 0:
                        filtered = dept_members
                    else:
                        target_label = all_job_labels[tab_idx - 1]
                        filtered = [
                            m for m in dept_members
                            if st.session_state.skill_members[m]["job_type_label"] == target_label
                        ]

                    sub1, sub2 = st.tabs(["📋 スキル項目（必須・選択）", "🔗 クロスファンクショナル"])
                    with sub1:
                        render_skill_comparison(
                            SKILL_ITEMS, "skill_scores", "prev_skill",
                            "skill_criteria", f"skill_{dept_key}_{tab_idx}",
                            filter_members=filtered,
                        )
                    with sub2:
                        render_skill_comparison(
                            CF_ITEMS, "cf_scores", "prev_cf",
                            "cf_criteria", f"cf_{dept_key}_{tab_idx}",
                            filter_members=filtered,
                        )

        if show_dept_tabs:
            dept_tab_labels = ["🏢 全部署"] + [f"📂 {d}" for d in sorted(dept_groups.keys())]
            dept_tabs = st.tabs(dept_tab_labels)

            for di, dept_tab in enumerate(dept_tabs):
                with dept_tab:
                    if di == 0:
                        dept_members = list(st.session_state.skill_members.keys())
                        dept_key = "all"
                    else:
                        dept_name = sorted(dept_groups.keys())[di - 1]
                        dept_members = dept_groups[dept_name]
                        dept_key = f"dept{di}"
                    render_job_tabs(dept_members, dept_key)
        else:
            render_job_tabs(list(st.session_state.skill_members.keys()), "all")

# ──────────────────────────────────────────────────────────────
# TAB 3: TARGET MANAGEMENT
# ──────────────────────────────────────────────────────────────
with t_target:
    st.header("目標管理評価シート")

    if not st.session_state.target_members:
        st.info("👈 まずアップロードタブで目標管理シートを登録してください。")
    else:
        members_t = list(st.session_state.target_members.keys())
        col_sel, col_mode = st.columns([2, 2])
        selected = col_sel.selectbox("👤 評価対象メンバー", members_t, key="target_select")
        view_mode = col_mode.radio(
            "表示モード",
            ["✏️ 下期UL評価入力", "📊 通期評価サマリー"],
            horizontal=True,
            key="target_mode",
        )

        if selected:
            d         = st.session_state.target_members[selected]
            grade_str = d.get("grade_str", "")
            job_type  = d.get("job_type", "非営業")
            weights   = get_weights(job_type, grade_str)
            weight_sum = sum(weights)

            st.info(
                f"**{selected}** ｜ 等級: **{grade_str}** ｜ 職種区分: **{job_type}** ｜ "
                f"ウェイト: 業績貢献={weights[0]} / 業績プロセス={weights[1]} / "
                f"会社貢献={weights[2]} / パーソナリティ={weights[3]}　合計={weight_sum}"
            )

            UL_OPTS = ["", "S", "A", "B", "C", "D"]

            # ════════════════════════════════════════════════
            # MODE A: 下期UL評価入力
            # ════════════════════════════════════════════════
            if view_mode == "✏️ 下期UL評価入力":
                total_score = 0.0

                for i, axis in enumerate(TARGET_AXES):
                    ax = d["axes"][axis]
                    w  = weights[i]

                    # ── カードヘッダー
                    st.markdown(
                        f'<div style="background:#f0f2f6;border-left:5px solid #4472C4;'
                        f'padding:8px 14px;border-radius:0 6px 0 0;margin-top:16px;">'
                        f'<strong style="font-size:1.05em;">{axis}</strong>'
                        f'<span style="color:#666;font-size:0.85em;margin-left:10px;">ウェイト: {w}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

                    # ── 目標・振り返り参照（折りたたみ）
                    lower_target = ax.get("lower_target", "")
                    lower_rev    = ax.get("lower_self_rev", "")
                    if lower_target or lower_rev:
                        with st.expander("📄 目標内容・本人振り返りを確認する", expanded=False):
                            ref_c1, ref_c2 = st.columns(2)
                            with ref_c1:
                                st.markdown("**📌 下期 目標設定**")
                                st.markdown(
                                    f'<div style="background:#fff8e1;border:1px solid #ffe082;'
                                    f'border-radius:6px;padding:10px;font-size:0.88em;'
                                    f'white-space:pre-wrap;max-height:200px;overflow-y:auto;">'
                                    f'{lower_target or "（記載なし）"}</div>',
                                    unsafe_allow_html=True,
                                )
                            with ref_c2:
                                st.markdown("**📝 本人 期末振り返り**")
                                st.markdown(
                                    f'<div style="background:#f3e5f5;border:1px solid #ce93d8;'
                                    f'border-radius:6px;padding:10px;font-size:0.88em;'
                                    f'white-space:pre-wrap;max-height:200px;overflow-y:auto;">'
                                    f'{lower_rev or "（記載なし）"}</div>',
                                    unsafe_allow_html=True,
                                )

                    # ── 評価入力行
                    inp_c1, inp_c2, inp_c3, inp_c4 = st.columns([1, 1, 1, 3])

                    self_eval = ax.get("self_eval", "") or "–"
                    inp_c1.metric("本人評価（参考）", self_eval)

                    curr_ul  = ax.get("ul_eval", "") or ""
                    curr_idx = UL_OPTS.index(curr_ul) if curr_ul in UL_OPTS else 0
                    new_ul   = inp_c2.selectbox(
                        "UL評価",
                        UL_OPTS,
                        index=curr_idx,
                        key=f"ul_{selected}_{axis}",
                    )

                    coeff = EVAL_COEFF.get(new_ul, 0.0)
                    score = round(w * coeff, 2)
                    total_score += score
                    inp_c3.metric("獲得スコア", score)

                    comment = inp_c4.text_area(
                        "UL評価コメント",
                        value=ax.get("ul_comment", "") or "",
                        key=f"comment_{selected}_{axis}",
                        height=90,
                        placeholder="評価コメントを入力...",
                    )

                    ax["ul_eval"]     = new_ul
                    ax["final_eval"]  = new_ul
                    ax["ul_total"]    = new_ul
                    ax["final_total"] = new_ul
                    ax["weight"]      = w
                    ax["score"]       = score
                    ax["ul_comment"]  = comment

                # ── 総合スコア
                rank = score_to_rank(total_score, weight_sum)
                rank_color = RANK_COLOR.get(rank, "#333")
                rate = round(total_score / weight_sum, 3) if weight_sum > 0 else 0

                st.markdown(
                    f'<div style="margin-top:20px;background:#f8f9fa;border:2px solid {rank_color};'
                    f'border-radius:8px;padding:14px 20px;">'
                    f'<span style="font-size:1.0em;">下期 獲得スコア合計: <strong>{round(total_score,2)}</strong>'
                    f' ／ {weight_sum}点　達成率: <strong>{round(rate*100,1)}%</strong>　'
                    f'目安評価: <strong style="color:{rank_color};font-size:1.3em;">{rank}</strong></span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                st.caption("※ 目安評価（11段階）: S=135%超 ／ A+=125%超 ／ A=115%超 ／ A－=105%超 ／ B+=102%超 ／ B=98%超〜102%以下 ／ B－=95%超 ／ C+=85%超 ／ C=75%超 ／ C－=65%超 ／ D=65%以下")

                st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
                if st.button("💾 評価を保存", key=f"save_target_{selected}", type="primary"):
                    for axis in TARGET_AXES:
                        ax = d["axes"][axis]
                        audit(selected, "目標管理", axis, "", ax["ul_eval"], ax.get("ul_comment",""))
                    st.success("✅ 評価を保存しました！")

            # ════════════════════════════════════════════════
            # MODE B: 通期評価サマリー
            # ════════════════════════════════════════════════
            else:
                st.markdown("#### 通期評価サマリー")
                st.caption("上期・下期・通期の評価を横断表示します。上期データはExcelから読み取った値です。")

                # ── 上期 / 下期 並列カード
                for i, axis in enumerate(TARGET_AXES):
                    ax = d["axes"][axis]
                    w  = weights[i]

                    st.markdown(
                        f'<div style="background:#f0f2f6;border-left:5px solid #4472C4;'
                        f'padding:8px 14px;border-radius:0 6px 0 0;margin-top:18px;">'
                        f'<strong>{axis}</strong>'
                        f'<span style="color:#666;font-size:0.85em;margin-left:10px;">'
                        f'ウェイト(通期): {w}</span></div>',
                        unsafe_allow_html=True,
                    )

                    col_upper, col_lower = st.columns(2)

                    with col_upper:
                        st.markdown(
                            '<div style="background:#E3F2FD;border-radius:0 0 0 6px;'
                            'padding:8px 12px;font-weight:bold;font-size:0.9em;">📅 上期</div>',
                            unsafe_allow_html=True,
                        )
                        u_target = ax.get("upper_target","") or ax.get("upper_jisseki","")
                        u_rev    = ax.get("upper_self_rev","")
                        u_ulcmt  = ax.get("upper_ul_cmt","")
                        u_self   = ax.get("upper_self_eval","") or "–"
                        u_ul     = ax.get("upper_ul_eval","") or "–"
                        u_final  = ax.get("upper_final","") or "–"
                        u_score  = ax.get("upper_score","") or "–"

                        if u_target:
                            with st.expander("目標設定・振り返りを見る", expanded=False):
                                st.markdown(f"**目標:** {u_target[:200]}")
                                if u_rev:
                                    st.markdown(f"**本人振り返り:** {u_rev[:200]}")
                                if u_ulcmt:
                                    st.markdown(f"**ULコメント:** {u_ulcmt[:200]}")

                        ec1,ec2,ec3,ec4 = st.columns(4)
                        ec1.metric("本人評価",u_self)
                        ec2.metric("UL評価",u_ul)
                        ec3.metric("最終評価",u_final)
                        ec4.metric("スコア",u_score)

                    with col_lower:
                        st.markdown(
                            '<div style="background:#E8F5E9;border-radius:0 0 6px 0;'
                            'padding:8px 12px;font-weight:bold;font-size:0.9em;">📅 下期</div>',
                            unsafe_allow_html=True,
                        )
                        l_target = ax.get("lower_target","")
                        l_rev    = ax.get("lower_self_rev","")
                        l_self   = ax.get("self_eval","") or "–"
                        l_ul     = ax.get("ul_eval","") or "–"
                        l_final  = ax.get("final_eval","") or "–"
                        l_score  = ax.get("score","") or "–"

                        if l_target:
                            with st.expander("目標設定・振り返りを見る", expanded=False):
                                st.markdown(f"**目標:** {l_target[:200]}")
                                if l_rev:
                                    st.markdown(f"**本人振り返り:** {l_rev[:200]}")
                                cmt = ax.get("ul_comment","")
                                if cmt:
                                    st.markdown(f"**ULコメント:** {cmt[:200]}")

                        dc1,dc2,dc3,dc4 = st.columns(4)
                        dc1.metric("本人評価",l_self)
                        dc2.metric("UL評価",l_ul)
                        dc3.metric("最終評価",l_final)
                        dc4.metric("スコア",l_score)

                    # 通期評価（入力可）
                    st.markdown(
                        '<div style="background:#FFF3E0;padding:6px 12px;'
                        'font-size:0.85em;font-weight:bold;">🔄 通期評価</div>',
                        unsafe_allow_html=True,
                    )
                    tc1,tc2,tc3 = st.columns([1,1,2])
                    curr_ut  = ax.get("ul_total","") or ""
                    curr_uti = UL_OPTS.index(curr_ut) if curr_ut in UL_OPTS else 0
                    new_ut = tc1.selectbox(
                        "通期 UL評価",
                        UL_OPTS, index=curr_uti,
                        key=f"ut_{selected}_{axis}",
                    )
                    ax["ul_total"] = new_ut

                    curr_ft  = ax.get("final_total","") or ""
                    curr_fti = UL_OPTS.index(curr_ft) if curr_ft in UL_OPTS else 0
                    new_ft = tc2.selectbox(
                        "通期 最終評価",
                        UL_OPTS, index=curr_fti,
                        key=f"ft_{selected}_{axis}",
                    )
                    ax["final_total"] = new_ft

                    coeff_t = EVAL_COEFF.get(new_ft or new_ut, 0.0)
                    score_t = round(w * coeff_t, 2)
                    ax["weight"] = w
                    ax["score"]  = score_t
                    tc3.metric("通期獲得スコア", score_t)

                # ── 通期総合
                total_annual = sum(
                    EVAL_COEFF.get(d["axes"][ax].get("final_total","") or
                                   d["axes"][ax].get("ul_total",""), 0) * weights[i]
                    for i, ax in enumerate(TARGET_AXES)
                )
                rank_a = score_to_rank(total_annual, weight_sum)
                rate_a = round(total_annual / weight_sum, 3) if weight_sum > 0 else 0
                rank_color_a = RANK_COLOR.get(rank_a, "#333")

                st.markdown(
                    f'<div style="margin-top:24px;background:#f8f9fa;border:2px solid {rank_color_a};'
                    f'border-radius:8px;padding:14px 20px;">'
                    f'通期 獲得スコア: <strong>{round(total_annual,2)}</strong> ／ {weight_sum}点　'
                    f'達成率: <strong>{round(rate_a*100,1)}%</strong>　'
                    f'目安評価: <strong style="color:{rank_color_a};font-size:1.3em;">{rank_a}</strong>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                if st.button("💾 通期評価を保存", key=f"save_annual_{selected}", type="primary"):
                    for axis in TARGET_AXES:
                        ax = d["axes"][axis]
                        audit(selected, "目標管理(通期)", axis, "", ax.get("final_total",""), "")
                    st.success("✅ 通期評価を保存しました！")

# ──────────────────────────────────────────────────────────────
# TAB 4: HISTORY
# ──────────────────────────────────────────────────────────────
with t_history:
    st.header("評価履歴")

    if not st.session_state.audit_log:
        st.info("まだ評価履歴がありません。評価を保存すると記録されます。")
    else:
        df_log = pd.DataFrame(st.session_state.audit_log)
        fc1, fc2 = st.columns(2)
        all_members_h = sorted(df_log["member"].unique().tolist())
        sel_m = fc1.multiselect("メンバー", all_members_h, default=all_members_h)
        sel_k = fc2.multiselect("種別", ["スキル", "目標管理"], default=["スキル", "目標管理"])

        filtered_log = df_log[df_log["member"].isin(sel_m) & df_log["kind"].isin(sel_k)]
        st.dataframe(
            filtered_log[["timestamp","member","kind","field","prev","new","comment"]].rename(
                columns={"timestamp":"日時","member":"メンバー","kind":"種別",
                         "field":"項目","prev":"変更前","new":"変更後","comment":"コメント"}
            ).sort_values("日時", ascending=False),
            use_container_width=True,
            hide_index=True,
        )
        st.download_button(
            "📥 履歴をCSVでダウンロード",
            data=filtered_log.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig"),
            file_name=f"評価履歴_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
        )

# ──────────────────────────────────────────────────────────────
# TAB 5: EXPORT
# ──────────────────────────────────────────────────────────────
with t_export:
    st.header("出力")
    ec1, ec2, ec3 = st.columns(3)

    with ec1:
        st.subheader("📊 スキルシート")
        n_skill = len(st.session_state.skill_members)
        st.caption(f"{n_skill}名分。変更箇所は赤字になります。")
        if n_skill == 0:
            st.info("スキルシートが未登録です。")
        else:
            if st.button("📦 ZIPを生成", key="gen_skill", use_container_width=True):
                zip_buf = BytesIO()
                with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for nm, d in st.session_state.skill_members.items():
                        try:
                            zf.writestr(f"スキルシート_{nm}.xlsx", export_skill_excel(d))
                        except Exception as e:
                            st.warning(f"⚠️ {nm}: {e}")
                zip_buf.seek(0)
                st.download_button(
                    "📥 ZIPダウンロード",
                    data=zip_buf,
                    file_name=f"スキルシート_{datetime.now().strftime('%Y%m%d')}.zip",
                    mime="application/zip",
                    key="dl_skill_zip",
                    use_container_width=True,
                )

    with ec2:
        st.subheader("📋 目標管理シート")
        n_target = len(st.session_state.target_members)
        st.caption(f"{n_target}名分。UL評価・コメント・スコアを書き込みます。")
        if n_target == 0:
            st.info("目標管理シートが未登録です。")
        else:
            if st.button("📦 ZIPを生成", key="gen_target", use_container_width=True):
                zip_buf = BytesIO()
                with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for nm, d in st.session_state.target_members.items():
                        try:
                            zf.writestr(f"目標管理評価_{nm}.xlsx", export_target_excel(d))
                        except Exception as e:
                            st.warning(f"⚠️ {nm}: {e}")
                zip_buf.seek(0)
                st.download_button(
                    "📥 ZIPダウンロード",
                    data=zip_buf,
                    file_name=f"目標管理評価_{datetime.now().strftime('%Y%m%d')}.zip",
                    mime="application/zip",
                    key="dl_target_zip",
                    use_container_width=True,
                )

    with ec3:
        st.subheader("📑 評価一覧")
        st.caption("全メンバーのサマリーをまとめたExcelを出力します。")
        total_m = len(set(
            list(st.session_state.skill_members.keys()) +
            list(st.session_state.target_members.keys())
        ))
        if total_m == 0:
            st.info("データが未登録です。")
        else:
            if st.button("📄 一覧Excelを生成", key="gen_summary", use_container_width=True):
                try:
                    st.download_button(
                        "📥 評価一覧をダウンロード",
                        data=export_summary_excel(),
                        file_name=f"評価一覧_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_summary",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"エラー: {e}")

    st.divider()
    st.caption("⚠️ セッションをリロードするとデータはリセットされます。評価が完了したら必ず出力してください。")
